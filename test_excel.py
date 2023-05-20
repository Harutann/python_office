from openpyxl import Workbook
wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])

import datetime
ws['A2'] = datetime.datetime.now()

ws_new = wb.create_sheet("追加シート")
value_data = 0
for x in range(1, 11):
    for y in range(1, 11):
        value_data = 10 * (x - 1) + y
        ws_new.cell(row=x, column=y).value = value_data

wb.save("sample.xlsx")